"""Reads the Excel workbook and returns typed domain objects for each sheet."""
from typing import List, Tuple
from datetime import date, datetime
import calendar
import openpyxl

import unicodedata

from domain.models import Eslabon, Medicamento, Parametro, Perfil, Printer, Usuario, StockRow, PerfilPermisoRaw
from config.settings import (
    HEADER_ROW, DATA_START_ROW,
    SHEET_REGISTRO, SHEET_PRODUCTOS, SHEET_PROVEEDORES, SHEET_PARAMETROS,
    SHEET_USUARIOS, SHEET_IMPRESORAS, SHEET_STOCK, SHEET_PERFIL_PERMISO,
    EMPRESA_LOOKUP, SPECIAL_QUOTED_PARAMS, DEFAULT_PASSWORD, PRINTER_TYPE_MAP,
)
from utils.logger import get_logger

logger = get_logger()


# ── helpers ──────────────────────────────────────────────────────────────────

def _cell_value(ws, row: int, col: int):
    """Return the cell value or None for empty / whitespace strings."""
    val = ws.cell(row=row, column=col).value
    if isinstance(val, str):
        val = val.strip() or None
    return val


def _to_str(val) -> str | None:
    return str(val) if val is not None else None


def _normalize_cuit(val) -> str | None:
    """Acepta CUIT con o sin guiones (ej: 11-11111111-1 o 11111111111) y devuelve solo dígitos."""
    s = _to_str(val)
    if s is None:
        return None
    return s.replace("-", "")


def _to_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _parse_fecha_vto(val, row_num: int) -> date:
    """
    Parse a date value that may come as:
      - A Python date/datetime object (openpyxl already converted it)
      - A string in dd/mm/yyyy, dd-mm-yyyy, or yyyy-mm-dd format
    """
    if val is None:
        raise ValueError(f"Fila {row_num}: 'Vencimiento' no puede estar vacío.")
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    raise ValueError(
        f"Fila {row_num}: formato de fecha no reconocido en 'Vencimiento': {val!r}. "
        f"Se aceptan dd/mm/yyyy, dd-mm-yyyy o yyyy-mm-dd."
    )


def _normalize_printer_type(s: str) -> str:
    """Uppercase and strip diacritics so 'Impresora Agrupación' == 'IMPRESORA AGRUPACION'."""
    return "".join(
        c for c in unicodedata.normalize("NFD", s.upper())
        if unicodedata.category(c) != "Mn"
    )


def _lookup_empresa(raw_value, row_num: int) -> Tuple[int | None, str | None]:
    """
    Resolve Tipo de Empresa text → ID_EMPRESA integer.
    Returns (id, error_message). error_message is None on success.
    """
    if raw_value is None:
        return None, f"Fila {row_num}: 'Tipo de Empresa' está vacío."
    normalised = str(raw_value).strip().upper()
    id_empresa = EMPRESA_LOOKUP.get(normalised)
    if id_empresa is None:
        return None, (
            f"Fila {row_num}: valor desconocido en 'Tipo de Empresa': '{raw_value}'. "
            f"Valores válidos: {list(EMPRESA_LOOKUP.keys())}"
        )
    return id_empresa, None


# ── sheet readers ─────────────────────────────────────────────────────────────

def read_formulario_registro(wb: openpyxl.Workbook) -> List[Eslabon]:
    """
    Sheet: 'Formulario de Registro' → list[Eslabon]

    Column layout (header row 9):
      A: Nombre eslabón  → RSOC
      B: GLN             → GLN
      C: CUIT            → CUIT
      D: Usuario ANMAT   → USER_ANMAT
      E: Contraseña ANMAT→ PASS_ANMAT
      F: URL             → URL
      G: Tipo de Empresa → ID_EMPRESA
    """
    ws = wb[SHEET_REGISTRO]
    eslabones: List[Eslabon] = []

    for row in range(DATA_START_ROW, ws.max_row + 1):
        rsoc = _cell_value(ws, row, 1)

        # Stop conditions: empty key column or "Responsables" section marker
        if rsoc is None:
            break
        if "responsables" in str(rsoc).lower():
            break

        gln        = _to_str(_cell_value(ws, row, 2))
        cuit       = _normalize_cuit(_cell_value(ws, row, 3))
        user_anmat = _to_str(_cell_value(ws, row, 4))
        pass_anmat = _to_str(_cell_value(ws, row, 5))
        url        = _to_str(_cell_value(ws, row, 6))
        tipo_raw   = _cell_value(ws, row, 7)

        id_empresa, err = _lookup_empresa(tipo_raw, row)
        if err:
            raise ValueError(f"[{SHEET_REGISTRO}] {err}")

        eslabones.append(Eslabon(
            RSOC=_to_str(rsoc),
            GLN=gln,
            CUIT=cuit,
            USER_ANMAT=user_anmat,
            PASS_ANMAT=pass_anmat,
            URL=url,
            ID_EMPRESA=id_empresa,
            ACTIVO=1,
        ))

    logger.debug(f"[{SHEET_REGISTRO}] {len(eslabones)} filas leídas.")
    return eslabones


def read_formulario_productos(wb: openpyxl.Workbook) -> List[Medicamento]:
    """
    Sheet: 'Formulario de Productos' → list[Medicamento]

    Column layout (header row 9):
      A: Nombre              → NOMBRE
      B: GTIN                → BC_EAN_1
      C: Código ERP          → BC_EAN_2
      D: Items por Pack      → ITEMS_POR_PACK
      E: Packs por Pallet    → PACKS_POR_PALLET
      F: Capas por Pack      → CANTIDAD_CAPAS_PACK
      G: Nivel de Agregación → LEVEL_AGGREGATION
    """
    ws = wb[SHEET_PRODUCTOS]
    medicamentos: List[Medicamento] = []

    for row in range(DATA_START_ROW, ws.max_row + 1):
        nombre = _cell_value(ws, row, 1)
        if nombre is None:
            break

        bc_ean_1           = _to_str(_cell_value(ws, row, 2))
        bc_ean_2           = _to_str(_cell_value(ws, row, 3))
        items_por_pack     = _to_int(_cell_value(ws, row, 4))
        packs_por_pallet   = _to_int(_cell_value(ws, row, 5))
        cantidad_capas_pack= _to_int(_cell_value(ws, row, 6))
        level_aggregation  = _to_int(_cell_value(ws, row, 7))

        medicamentos.append(Medicamento(
            NOMBRE=_to_str(nombre),
            BC_EAN_1=bc_ean_1,
            BC_EAN_2=bc_ean_2,
            ITEMS_POR_PACK=items_por_pack,
            PACKS_POR_PALLET=packs_por_pallet,
            ID_LAB=1,
            TRAZABLE=1,
            EXIGIBLE=1,
            ACTIVO=1,
            SERIE=1,
            CANTIDAD_CAPAS_PACK=cantidad_capas_pack,
            LEVEL_AGGREGATION=level_aggregation,
        ))

    logger.debug(f"[{SHEET_PRODUCTOS}] {len(medicamentos)} filas leídas.")
    return medicamentos


def read_formulario_proveedores(wb: openpyxl.Workbook) -> List[Eslabon]:
    """
    Sheet: 'Formulario Proveedores Clientes' → list[Eslabon]

    Column layout (header row 9):
      A: Razón Social    → RSOC
      B: Dirección       → DIRECCION
      C: GLN             → GLN
      D: CUIT            → CUIT
      E: Tipo de Empresa → ID_EMPRESA
    """
    ws = wb[SHEET_PROVEEDORES]
    eslabones: List[Eslabon] = []

    for row in range(DATA_START_ROW, ws.max_row + 1):
        rsoc = _cell_value(ws, row, 1)
        if rsoc is None:
            break

        direccion  = _to_str(_cell_value(ws, row, 2))
        gln        = _to_str(_cell_value(ws, row, 3))
        cuit       = _normalize_cuit(_cell_value(ws, row, 4))
        tipo_raw   = _cell_value(ws, row, 5)

        id_empresa, err = _lookup_empresa(tipo_raw, row)
        if err:
            raise ValueError(f"[{SHEET_PROVEEDORES}] {err}")

        eslabones.append(Eslabon(
            RSOC=_to_str(rsoc),
            GLN=gln,
            CUIT=cuit,
            DIRECCION=direccion,
            ID_EMPRESA=id_empresa,
            ACTIVO=1,
        ))

    logger.debug(f"[{SHEET_PROVEEDORES}] {len(eslabones)} filas leídas.")
    return eslabones


def _detect_columns(ws, header_row: int, targets: dict) -> dict:
    """
    Scan the header row and return a mapping of {key: col_index (1-based)}.

    targets = {"nombre": ["configuración", "configuracion"], "valor": ["valor"]}
    Matching is case-insensitive and ignores leading/trailing spaces.
    Raises ValueError if a required column is not found.
    """
    found = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=header_row, column=col).value
        if raw is None:
            continue
        header_text = str(raw).strip().lower()
        for key, aliases in targets.items():
            if key not in found and header_text in aliases:
                found[key] = col

    missing = [k for k in targets if k not in found]
    if missing:
        raise ValueError(
            f"[{SHEET_PARAMETROS}] No se encontraron las columnas de cabecera: {missing}. "
            f"Verificar que la fila {header_row} contiene los encabezados correctos."
        )
    return found


def read_parametros(wb: openpyxl.Workbook) -> List[Parametro]:
    """
    Sheet: 'Parámetros' → list[Parametro]

    Columns are detected dynamically from the header row (row 9) so that
    variations in column order or extra columns in the Excel don't break parsing.

    Expected headers (case-insensitive):
      'Configuración' or 'Configuracion' → NOMBRE  (lookup key)
      'Valor'                            → VALOR
      'Descripción'                      → ignored
    """
    ws = wb[SHEET_PARAMETROS]
    parametros: List[Parametro] = []

    # Detect column positions from the actual header row
    col_map = _detect_columns(ws, HEADER_ROW, {
        "nombre": ["configuración", "configuracion", "configuración:", "configuracion:"],
        "valor":  ["valor", "valor:"],
    })
    col_nombre = col_map["nombre"]
    col_valor  = col_map["valor"]
    logger.debug(
        f"[{SHEET_PARAMETROS}] Columnas detectadas → "
        f"NOMBRE=col{col_nombre}, VALOR=col{col_valor}"
    )

    for row in range(DATA_START_ROW, ws.max_row + 1):
        nombre = _cell_value(ws, row, col_nombre)
        if nombre is None:
            break

        raw_valor  = _cell_value(ws, row, col_valor)
        nombre_str = str(nombre).strip()

        # Build final value
        valor_str = str(raw_valor).strip() if raw_valor is not None else ""

        # Normalize SI/NO → TRUE/FALSE (case-insensitive)
        valor_upper = valor_str.upper()
        if valor_upper == "SI":
            valor_str = "TRUE"
        elif valor_upper == "NO":
            valor_str = "FALSE"

        if nombre_str in SPECIAL_QUOTED_PARAMS:
            valor_str = f"'{valor_str}'"

        parametros.append(Parametro(NOMBRE=nombre_str, VALOR=valor_str))

    logger.debug(f"[{SHEET_PARAMETROS}] {len(parametros)} filas leídas.")
    return parametros


def read_formulario_usuarios(wb: openpyxl.Workbook) -> List[Usuario]:
    """
    Sheet: 'Formulario de Usuarios' → list[Usuario]

    Column layout (header row 9):
      A: Nombre y Apellido:              → APEYNOM       (nullable)
      B: Usuario:                        → USERNAME       (required)
      C: Contraseña (por defecto ...):   → PASSWORD_RAW   (empty → DEFAULT_PASSWORD)
      D: Perfil:                         → ID_PERFIL      (resolved later via DB lookup)
      E: Posición/Cargo:                 → CARGO          (nullable)
      F: Mail:                           → MAIL           (nullable)
      G: Bonificado (Si / No):           → FACTURA + FECHA_VTO
      H: URL acceso:                     → ID_ESLABON     (resolved later via DB lookup)

    Note: ID_PERFIL and ID_ESLABON are stored as raw strings here and resolved
    by the import service after DB queries.
    """
    ws = wb[SHEET_USUARIOS]
    usuarios: List[Usuario] = []
    today = date.today()

    for row in range(DATA_START_ROW, ws.max_row + 1):
        username = _cell_value(ws, row, 2)
        if username is None:
            break

        apeynom      = _to_str(_cell_value(ws, row, 1))
        password_raw = _to_str(_cell_value(ws, row, 3)) or DEFAULT_PASSWORD
        perfil_nombre= _to_str(_cell_value(ws, row, 4))
        cargo        = _to_str(_cell_value(ws, row, 5))
        mail         = _to_str(_cell_value(ws, row, 6))
        bonificado   = _to_str(_cell_value(ws, row, 7))
        url_acceso   = _to_str(_cell_value(ws, row, 8))

        # Validate required fields
        if not perfil_nombre:
            raise ValueError(
                f"[{SHEET_USUARIOS}] Fila {row}: 'Perfil' no puede estar vacío."
            )
        if not url_acceso:
            raise ValueError(
                f"[{SHEET_USUARIOS}] Fila {row}: 'URL acceso' no puede estar vacío."
            )
        if not bonificado:
            raise ValueError(
                f"[{SHEET_USUARIOS}] Fila {row}: 'Bonificado' no puede estar vacío."
            )

        # Bonificado logic
        bonificado_upper = bonificado.strip().upper()
        if bonificado_upper == "SI":
            factura   = "BONIFICADO"
            fecha_vto = None
        elif bonificado_upper == "NO":
            factura   = "pendiente"
            fecha_vto = date(today.year + 1, today.month, today.day)
        else:
            raise ValueError(
                f"[{SHEET_USUARIOS}] Fila {row}: 'Bonificado' debe ser 'Si' o 'No', "
                f"recibido: '{bonificado}'"
            )

        # Store perfil_nombre and url_acceso as sentinel strings — resolved by service
        usuarios.append(Usuario(
            APEYNOM=apeynom,
            USERNAME=_to_str(username),
            PASSWORD_RAW=password_raw,
            ID_PERFIL=perfil_nombre,    # type: ignore — resolved to int by service
            CARGO=cargo,
            MAIL=mail,
            FACTURA=factura,
            FECHA_VTO=fecha_vto,
            ID_ESLABON=url_acceso,      # type: ignore — resolved to int by service
            ACTIVO=1,
            FECHA_ALTA=today,
        ))

    logger.debug(f"[{SHEET_USUARIOS}] {len(usuarios)} filas leídas.")
    return usuarios


def read_formulario_impresoras(wb: openpyxl.Workbook) -> List[Printer]:
    """
    Sheet: 'Formulario de Impresoras' → list[Printer]

    Column layout (header row 9):
      A: Nombre de Impresora:  → printer.nombre  (also determines all fixed fields)
      B: Tipo de Impresora:    → ignored (referencia humana, no mapeado a BD)
      C: Tipo de Conexión:     → printer.HOST  (dirección IP)
      D: Cantidad:             → fila omitida si Cantidad == 0

    El tipo de impresora se resuelve desde PRINTER_TYPE_MAP en config/settings.py.
    """
    ws = wb[SHEET_IMPRESORAS]
    printers: List[Printer] = []

    for row in range(DATA_START_ROW, ws.max_row + 1):
        nombre = _cell_value(ws, row, 1)
        if nombre is None:
            break

        cantidad_raw = _cell_value(ws, row, 4)
        try:
            cantidad = int(cantidad_raw) if cantidad_raw is not None else 0
        except (ValueError, TypeError):
            cantidad = 0

        if cantidad == 0:
            logger.debug(f"  [{SHEET_IMPRESORAS}] Fila {row}: '{nombre}' omitida (Cantidad=0).")
            continue

        tipo_key = _normalize_printer_type(str(nombre))
        cfg = PRINTER_TYPE_MAP.get(tipo_key)
        if cfg is None:
            raise ValueError(
                f"[{SHEET_IMPRESORAS}] Fila {row}: tipo de impresora desconocido: '{nombre}'. "
                f"Valores válidos (normalizados): {list(PRINTER_TYPE_MAP.keys())}"
            )

        ip = _to_str(_cell_value(ws, row, 3))
        host = ip if cfg["host_from_excel"] else None

        printers.append(Printer(
            nombre=_to_str(nombre),
            print_max=cfg["print_max"],
            tipo=cfg["tipo"],
            genera_packs=cfg["genera_packs"],
            genera_pallets=cfg["genera_pallets"],
            linea_produccion=cfg["linea_produccion"],
            id_pattern=cfg["id_pattern"],
            plugin=cfg["plugin"],
            klass=cfg["klass"],
            host=host,
            port=9999,
            configuracion_updates=list(cfg["configuracion_updates"]),
        ))

    logger.debug(f"[{SHEET_IMPRESORAS}] {len(printers)} impresoras activas leídas.")
    return printers


def read_formulario_stock(wb: openpyxl.Workbook) -> List[StockRow]:
    """
    Sheet: 'Formulario de Stock' → list[StockRow]

    Usa iter_rows() para soportar volúmenes grandes (ej. 100 000 filas) sin
    cargar toda la hoja en memoria.

    Columnas detectadas dinámicamente desde la fila de cabecera (HEADER_ROW):
      Lote           → cod_lote
      GTIN           → gtin
      Vencimiento    → fecha_vto   (dd/mm/yyyy | dd-mm-yyyy | yyyy-mm-dd)
      Serie          → serie
      Código Pack    → id_pack     ('0' si vacío)
      Código Pallet  → id_pallet   ('0' si vacío)
      URL acceso     → url_acceso
    """
    ws = wb[SHEET_STOCK]

    # ── detectar columnas desde la fila de cabecera ───────────────────────────
    ALIASES: dict[str, list[str]] = {
        "lote":   ["lote", "lote:"],
        "gtin":   ["gtin", "gtin:"],
        "vto":    ["vencimiento", "vencimiento:", "fecha vto", "fecha_vto"],
        "serie":  ["serie", "serie:"],
        "pack":   ["código pack", "codigo pack", "código pack:", "codigo pack:"],
        "pallet": ["código pallet", "codigo pallet", "código pallet:", "codigo pallet:"],
        "url":    ["url acceso", "url acceso:"],
    }

    header_map: dict[str, int] = {}  # normalized_header → 0-based col index
    for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True):
        for idx, cell_val in enumerate(row):
            if cell_val is not None:
                header_map[str(cell_val).strip().lower()] = idx
        break

    col_map: dict[str, int] = {}
    for key, aliases in ALIASES.items():
        for alias in aliases:
            if alias in header_map:
                col_map[key] = header_map[alias]
                break
        if key not in col_map:
            raise ValueError(
                f"[{SHEET_STOCK}] Columna no encontrada: '{key}'. "
                f"Aliases buscados: {aliases}. "
                f"Cabeceras detectadas: {list(header_map.keys())}"
            )

    logger.debug(f"[{SHEET_STOCK}] Columnas detectadas: {col_map}")

    # ── leer filas de datos via streaming ─────────────────────────────────────
    rows: List[StockRow] = []
    row_num = DATA_START_ROW - 1

    for row_values in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        row_num += 1

        def _get(key):
            idx = col_map[key]
            return row_values[idx] if idx < len(row_values) else None

        cod_lote_raw = _get("lote")
        if cod_lote_raw is None or str(cod_lote_raw).strip() == "":
            break
        cod_lote = str(cod_lote_raw).strip()

        gtin_raw = _get("gtin")
        gtin     = str(gtin_raw).strip() if gtin_raw is not None else ""

        serie_raw = _get("serie")
        serie     = str(serie_raw).strip() if serie_raw is not None else ""

        pack_raw  = _get("pack")
        id_pack   = str(pack_raw).strip() if pack_raw is not None and str(pack_raw).strip() else "0"

        pallet_raw = _get("pallet")
        id_pallet  = str(pallet_raw).strip() if pallet_raw is not None and str(pallet_raw).strip() else "0"

        url_raw    = _get("url")
        url_acceso = str(url_raw).strip() if url_raw is not None else ""

        fecha_vto = _parse_fecha_vto(_get("vto"), row_num)

        rows.append(StockRow(
            cod_lote=cod_lote,
            gtin=gtin,
            fecha_vto=fecha_vto,
            serie=serie,
            id_pack=id_pack,
            id_pallet=id_pallet,
            url_acceso=url_acceso,
        ))

    logger.debug(f"[{SHEET_STOCK}] {len(rows)} filas leídas.")
    return rows


def read_perfil_permiso(wb: openpyxl.Workbook) -> List[PerfilPermisoRaw]:
    """
    Sheet: 'Perfil Permiso' → list[PerfilPermisoRaw]

    Columnas detectadas dinámicamente desde la fila de cabecera (HEADER_ROW):
      Nombre de perfil  → nombre_perfil
      Acción:           → accion
      Módulo:           → modulo
      Script:           → script
    """
    ws = wb[SHEET_PERFIL_PERMISO]

    ALIASES: dict[str, list[str]] = {
        "perfil": ["nombre de perfil", "nombre de perfil:"],
        "accion": ["acción:", "accion:", "acción", "accion"],
        "modulo": ["módulo:", "modulo:", "módulo", "modulo"],
        "script": ["script:", "script"],
    }

    header_map: dict[str, int] = {}
    for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True):
        for idx, cell_val in enumerate(row):
            if cell_val is not None:
                header_map[str(cell_val).strip().lower()] = idx
        break

    col_map: dict[str, int] = {}
    for key, aliases in ALIASES.items():
        for alias in aliases:
            if alias in header_map:
                col_map[key] = header_map[alias]
                break
        if key not in col_map:
            raise ValueError(
                f"[{SHEET_PERFIL_PERMISO}] Columna no encontrada: '{key}'. "
                f"Aliases buscados: {aliases}. "
                f"Cabeceras detectadas: {list(header_map.keys())}"
            )

    logger.debug(f"[{SHEET_PERFIL_PERMISO}] Columnas detectadas: {col_map}")

    rows: List[PerfilPermisoRaw] = []
    for row_values in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        def _get(key):
            idx = col_map[key]
            val = row_values[idx] if idx < len(row_values) else None
            return str(val).strip() if val is not None and str(val).strip() else None

        nombre_perfil = _get("perfil")
        if nombre_perfil is None:
            break

        accion = _get("accion")
        if not accion:
            raise ValueError(
                f"[{SHEET_PERFIL_PERMISO}] 'Acción' no puede estar vacío "
                f"para el perfil '{nombre_perfil}'."
            )

        rows.append(PerfilPermisoRaw(
            nombre_perfil=nombre_perfil,
            accion=accion,
            modulo=_get("modulo"),
            script=_get("script"),
        ))

    logger.debug(f"[{SHEET_PERFIL_PERMISO}] {len(rows)} filas leídas.")
    return rows


# ── public entry point ────────────────────────────────────────────────────────

def read_workbook(path: str):
    """
    Open the workbook and return a dict of parsed domain objects.

    Returns:
        {
            "registro":    list[Eslabon],
            "productos":   list[Medicamento],
            "proveedores": list[Eslabon],
            "parametros":  list[Parametro],
        }
    """
    logger.info(f"Abriendo fichero Excel: {path}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    available = set(wb.sheetnames)
    for required in (
        SHEET_REGISTRO, SHEET_PRODUCTOS, SHEET_PROVEEDORES,
        SHEET_USUARIOS, SHEET_PARAMETROS, SHEET_IMPRESORAS,
    ):
        if required not in available:
            raise FileNotFoundError(
                f"La pestaña requerida '{required}' no se encontró en el fichero. "
                f"Pestañas disponibles: {list(available)}"
            )

    result = {
        "registro":    read_formulario_registro(wb),
        "productos":   read_formulario_productos(wb),
        "proveedores": read_formulario_proveedores(wb),
        "usuarios":    read_formulario_usuarios(wb),
        "parametros":  read_parametros(wb),
        "impresoras":  read_formulario_impresoras(wb),
        "stock":       [],
    }

    if SHEET_STOCK in available:
        result["stock"] = read_formulario_stock(wb)
    else:
        logger.info(f"Pestaña '{SHEET_STOCK}' no encontrada — omitida.")

    result["perfil_permiso"] = []
    if SHEET_PERFIL_PERMISO in available:
        result["perfil_permiso"] = read_perfil_permiso(wb)
    else:
        logger.info(f"Pestaña '{SHEET_PERFIL_PERMISO}' no encontrada — omitida.")

    return result
